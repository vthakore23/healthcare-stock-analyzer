import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import yfinance as yf
from datetime import datetime, timedelta
import io
import base64

# Page configuration
st.set_page_config(
    page_title="DCF Model Builder - Healthcare Analyzer | June 2025",
    page_icon="📊",
    layout="wide"
)

# Professional CSS styling
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    .dcf-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin-bottom: 2rem;
        font-family: 'Inter', sans-serif;
    }
    
    .assumption-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        border: 1px solid #e2e8f0;
    }
    
    .valuation-result {
        background: linear-gradient(135deg, #00C851 0%, #007E33 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        text-align: center;
        margin: 1.5rem 0;
        font-family: 'Inter', sans-serif;
    }
    
    .scenario-analysis {
        background: #f8fafc;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border: 1px solid #e2e8f0;
    }
    
    .metric-highlight {
        font-size: 2.5rem;
        font-weight: bold;
        color: white;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .assumption-input {
        background: #f8fafc;
        border: 2px solid #e2e8f0;
        border-radius: 8px;
        padding: 0.5rem;
        font-size: 1rem;
        width: 100%;
    }
    
    .sensitivity-grid {
        background: white;
        border-radius: 12px;
        padding: 1rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    }
</style>
""", unsafe_allow_html=True)

class DCFModel:
    """Comprehensive DCF (Discounted Cash Flow) Model Builder"""
    
    def __init__(self):
        self.projection_years = 10
        self.terminal_growth_rate = 2.5
        self.tax_rate = 21.0
        
    def calculate_dcf_valuation(self, assumptions):
        """Calculate DCF valuation based on user assumptions"""
        
        years = list(range(1, self.projection_years + 1))
        
        # Base year metrics
        base_revenue = assumptions['base_revenue']
        base_ebitda_margin = assumptions['base_ebitda_margin']
        
        # Growth rates
        revenue_growth_rates = self._generate_growth_rates(
            assumptions['revenue_growth_y1'],
            assumptions['revenue_growth_y5'],
            assumptions['revenue_growth_terminal']
        )
        
        # Build financial projections
        projections = self._build_projections(assumptions, revenue_growth_rates)
        
        # Calculate valuation
        valuation_results = self._calculate_valuation(projections, assumptions)
        
        return projections, valuation_results
    
    def _generate_growth_rates(self, y1_growth, y5_growth, terminal_growth):
        """Generate smooth revenue growth rates from Y1 to terminal"""
        
        growth_rates = []
        
        # Years 1-5: Linear decline from Y1 to Y5
        for year in range(1, 6):
            rate = y1_growth - (y1_growth - y5_growth) * ((year - 1) / 4)
            growth_rates.append(rate / 100)
        
        # Years 6-10: Linear decline from Y5 to terminal
        for year in range(6, 11):
            rate = y5_growth - (y5_growth - terminal_growth) * ((year - 5) / 5)
            growth_rates.append(rate / 100)
        
        return growth_rates
    
    def _build_projections(self, assumptions, growth_rates):
        """Build detailed financial projections"""
        
        projections = {
            'Year': list(range(1, self.projection_years + 1)),
            'Revenue': [],
            'Revenue_Growth': [],
            'EBITDA': [],
            'EBITDA_Margin': [],
            'EBIT': [],
            'Taxes': [],
            'NOPAT': [],
            'Capex': [],
            'Capex_Revenue': [],
            'Change_NWC': [],
            'FCFF': [],
            'Discount_Factor': [],
            'PV_FCFF': []
        }
        
        # Base year values
        current_revenue = assumptions['base_revenue']
        wacc = assumptions['wacc'] / 100
        
        for i, year in enumerate(range(1, self.projection_years + 1)):
            # Revenue
            growth_rate = growth_rates[i]
            revenue = current_revenue * (1 + growth_rate)
            projections['Revenue'].append(revenue)
            projections['Revenue_Growth'].append(growth_rate * 100)
            
            # EBITDA with margin improvement/decline
            ebitda_margin = assumptions['base_ebitda_margin'] + \
                           (assumptions['target_ebitda_margin'] - assumptions['base_ebitda_margin']) * \
                           min(year / 5, 1.0)  # Reach target by Year 5
            ebitda = revenue * (ebitda_margin / 100)
            projections['EBITDA'].append(ebitda)
            projections['EBITDA_Margin'].append(ebitda_margin)
            
            # Depreciation & Amortization
            da = revenue * (assumptions['da_revenue'] / 100)
            ebit = ebitda - da
            projections['EBIT'].append(ebit)
            
            # Taxes
            taxes = ebit * (self.tax_rate / 100) if ebit > 0 else 0
            projections['Taxes'].append(taxes)
            
            # NOPAT (Net Operating Profit After Tax)
            nopat = ebit - taxes
            projections['NOPAT'].append(nopat)
            
            # Capital Expenditures
            capex = revenue * (assumptions['capex_revenue'] / 100)
            projections['Capex'].append(capex)
            projections['Capex_Revenue'].append(assumptions['capex_revenue'])
            
            # Change in Net Working Capital
            nwc_change = (revenue - current_revenue) * (assumptions['nwc_revenue'] / 100)
            projections['Change_NWC'].append(nwc_change)
            
            # Free Cash Flow to the Firm
            fcff = nopat + da - capex - nwc_change
            projections['FCFF'].append(fcff)
            
            # Discount factors and present value
            discount_factor = 1 / ((1 + wacc) ** year)
            projections['Discount_Factor'].append(discount_factor)
            
            pv_fcff = fcff * discount_factor
            projections['PV_FCFF'].append(pv_fcff)
            
            # Update for next iteration
            current_revenue = revenue
        
        return pd.DataFrame(projections)
    
    def _calculate_valuation(self, projections, assumptions):
        """Calculate enterprise and equity valuation"""
        
        # Terminal value calculation
        terminal_fcff = projections['FCFF'].iloc[-1] * (1 + self.terminal_growth_rate / 100)
        terminal_value = terminal_fcff / (assumptions['wacc'] / 100 - self.terminal_growth_rate / 100)
        
        # Present value of terminal value
        pv_terminal_value = terminal_value / ((1 + assumptions['wacc'] / 100) ** self.projection_years)
        
        # Enterprise value
        pv_projection_period = projections['PV_FCFF'].sum()
        enterprise_value = pv_projection_period + pv_terminal_value
        
        # Equity value
        equity_value = enterprise_value - assumptions['net_debt']
        
        # Per share value
        value_per_share = equity_value / assumptions['shares_outstanding']
        
        # Current stock price comparison
        current_price = assumptions.get('current_stock_price', 0)
        upside_downside = ((value_per_share - current_price) / current_price * 100) if current_price > 0 else 0
        
        return {
            'pv_projection_period': pv_projection_period,
            'terminal_value': terminal_value,
            'pv_terminal_value': pv_terminal_value,
            'enterprise_value': enterprise_value,
            'equity_value': equity_value,
            'value_per_share': value_per_share,
            'current_price': current_price,
            'upside_downside': upside_downside,
            'implied_ev_revenue': enterprise_value / assumptions['base_revenue'],
            'implied_pe': equity_value / (projections['NOPAT'].iloc[0] * assumptions['shares_outstanding'] / equity_value) if projections['NOPAT'].iloc[0] > 0 else 0
        }
    
    def sensitivity_analysis(self, base_assumptions, sensitivity_vars):
        """Perform sensitivity analysis on key variables"""
        
        base_valuation = self.calculate_dcf_valuation(base_assumptions)[1]['value_per_share']
        
        sensitivity_results = {}
        
        for var_name, var_range in sensitivity_vars.items():
            results = []
            
            for var_value in var_range:
                # Create modified assumptions
                modified_assumptions = base_assumptions.copy()
                modified_assumptions[var_name] = var_value
                
                # Calculate valuation
                try:
                    _, valuation = self.calculate_dcf_valuation(modified_assumptions)
                    results.append({
                        'variable_value': var_value,
                        'value_per_share': valuation['value_per_share'],
                        'change_vs_base': (valuation['value_per_share'] - base_valuation) / base_valuation * 100
                    })
                except:
                    results.append({
                        'variable_value': var_value,
                        'value_per_share': 0,
                        'change_vs_base': -100
                    })
            
            sensitivity_results[var_name] = pd.DataFrame(results)
        
        return sensitivity_results
    
    def scenario_analysis(self, base_assumptions):
        """Perform scenario analysis (Bull, Base, Bear cases)"""
        
        scenarios = {
            'Bear Case': {
                'revenue_growth_y1': base_assumptions['revenue_growth_y1'] - 5,
                'revenue_growth_y5': base_assumptions['revenue_growth_y5'] - 3,
                'target_ebitda_margin': base_assumptions['target_ebitda_margin'] - 2,
                'wacc': base_assumptions['wacc'] + 1,
                'terminal_growth_rate': 1.5
            },
            'Base Case': base_assumptions.copy(),
            'Bull Case': {
                'revenue_growth_y1': base_assumptions['revenue_growth_y1'] + 5,
                'revenue_growth_y5': base_assumptions['revenue_growth_y5'] + 3,
                'target_ebitda_margin': base_assumptions['target_ebitda_margin'] + 2,
                'wacc': base_assumptions['wacc'] - 0.5,
                'terminal_growth_rate': 3.0
            }
        }
        
        scenario_results = {}
        
        for scenario_name, scenario_assumptions in scenarios.items():
            # Merge with base assumptions
            full_assumptions = base_assumptions.copy()
            full_assumptions.update(scenario_assumptions)
            
            try:
                projections, valuation = self.calculate_dcf_valuation(full_assumptions)
                scenario_results[scenario_name] = {
                    'projections': projections,
                    'valuation': valuation
                }
            except Exception as e:
                scenario_results[scenario_name] = {
                    'projections': None,
                    'valuation': {'value_per_share': 0, 'enterprise_value': 0}
                }
        
        return scenario_results

class FinancialDataScraper:
    """Enhanced financial data scraper for DCF modeling"""
    
    def __init__(self):
        self.cache = {}
    
    def get_comprehensive_data(self, ticker):
        """Get comprehensive financial data for DCF modeling"""
        
        if ticker in self.cache:
            return self.cache[ticker]
        
        try:
            stock = yf.Ticker(ticker)
            
            # Basic company info
            info = stock.info
            
            # Financial statements
            financials = stock.financials
            balance_sheet = stock.balance_sheet
            cash_flow = stock.cashflow
            
            # Historical data for growth calculations
            hist_data = stock.history(period="5y")
            
            # Process and structure the data
            company_data = self._process_financial_data(
                info, financials, balance_sheet, cash_flow, hist_data, ticker
            )
            
            # Cache the results
            self.cache[ticker] = company_data
            
            return company_data
            
        except Exception as e:
            st.error(f"❌ Error fetching data for {ticker}: {e}")
            return None
    
    def _process_financial_data(self, info, financials, balance_sheet, cash_flow, hist_data, ticker):
        """Process raw financial data into DCF-ready format"""
        
        try:
            # Get most recent annual data (TTM)
            latest_revenue = self._get_latest_value(financials, 'Total Revenue', default=0)
            latest_ebitda = self._get_latest_value(financials, 'EBITDA', default=0)
            latest_ebit = self._get_latest_value(financials, 'EBIT', default=0)
            
            # Balance sheet items
            total_debt = self._get_latest_value(balance_sheet, 'Total Debt', default=0)
            total_cash = self._get_latest_value(balance_sheet, 'Cash And Cash Equivalents', default=0)
            
            # Cash flow items
            operating_cf = self._get_latest_value(cash_flow, 'Operating Cash Flow', default=0)
            capex = abs(self._get_latest_value(cash_flow, 'Capital Expenditures', default=0))
            
            # Calculate key metrics
            ebitda_margin = (latest_ebitda / latest_revenue * 100) if latest_revenue > 0 else 0
            net_debt = (total_debt - total_cash) / 1e6  # Convert to millions
            
            # Historical growth rates
            revenue_growth_3y = self._calculate_cagr(financials, 'Total Revenue', 3)
            revenue_growth_5y = self._calculate_cagr(financials, 'Total Revenue', 5)
            
            # Free cash flow
            free_cash_flow = operating_cf - capex if operating_cf and capex else 0
            
            # Shares outstanding and market data
            shares_outstanding = info.get('sharesOutstanding', info.get('impliedSharesOutstanding', 0))
            current_price = info.get('regularMarketPrice', info.get('currentPrice', 0))
            
            # Industry and beta for WACC estimation
            beta = info.get('beta', 1.0)
            industry = info.get('industry', 'Healthcare')
            
            return {
                'company_info': {
                    'name': info.get('longName', ticker),
                    'ticker': ticker,
                    'industry': industry,
                    'sector': info.get('sector', 'Healthcare'),
                    'country': info.get('country', 'US')
                },
                'financial_metrics': {
                    'revenue_ttm': latest_revenue / 1e6,  # Convert to millions
                    'ebitda_ttm': latest_ebitda / 1e6,
                    'ebit_ttm': latest_ebit / 1e6,
                    'operating_cf_ttm': operating_cf / 1e6,
                    'free_cf_ttm': free_cash_flow / 1e6,
                    'capex_ttm': capex / 1e6,
                    'ebitda_margin': ebitda_margin,
                    'capex_revenue_ratio': (capex / latest_revenue * 100) if latest_revenue > 0 else 3.0
                },
                'balance_sheet': {
                    'total_debt': total_debt / 1e6,
                    'total_cash': total_cash / 1e6,
                    'net_debt': net_debt
                },
                'market_data': {
                    'shares_outstanding': shares_outstanding / 1e6,  # Convert to millions
                    'current_price': current_price,
                    'market_cap': (shares_outstanding * current_price) / 1e6,
                    'beta': beta
                },
                'growth_metrics': {
                    'revenue_cagr_3y': revenue_growth_3y,
                    'revenue_cagr_5y': revenue_growth_5y,
                    'estimated_wacc': self._estimate_wacc(beta, info)
                }
            }
            
        except Exception as e:
            st.error(f"Error processing financial data: {e}")
            return None
    
    def _get_latest_value(self, df, column_name, default=0):
        """Get the most recent value from financial statement DataFrame"""
        try:
            if df is not None and not df.empty and column_name in df.index:
                # Get the most recent non-null value
                series = df.loc[column_name]
                for value in series:
                    if pd.notna(value) and value != 0:
                        return float(value)
            return default
        except:
            return default
    
    def _calculate_cagr(self, df, metric_name, years):
        """Calculate Compound Annual Growth Rate"""
        try:
            if df is not None and not df.empty and metric_name in df.index:
                series = df.loc[metric_name]
                values = [v for v in series if pd.notna(v) and v > 0]
                
                if len(values) >= 2:
                    # Get the most recent and oldest values
                    latest = values[0]
                    oldest = values[-1] if len(values) >= years else values[-1]
                    periods = min(len(values) - 1, years - 1)
                    
                    if periods > 0 and oldest > 0:
                        cagr = ((latest / oldest) ** (1/periods) - 1) * 100
                        return cagr
            return 0
        except:
            return 0
    
    def _estimate_wacc(self, beta, info):
        """Estimate WACC based on company characteristics"""
        
        # Risk-free rate (approximate current 10-year Treasury)
        risk_free_rate = 4.5
        
        # Market risk premium
        market_risk_premium = 6.0
        
        # Cost of equity using CAPM
        cost_of_equity = risk_free_rate + (beta * market_risk_premium)
        
        # Adjust for company-specific factors
        if info.get('marketCap', 0) < 2e9:  # Small cap
            cost_of_equity += 2.0
        elif info.get('marketCap', 0) < 10e9:  # Mid cap
            cost_of_equity += 1.0
        
        # Healthcare/biotech adjustment
        sector = info.get('sector', '').lower()
        if 'healthcare' in sector or 'biotech' in sector:
            cost_of_equity += 1.0  # Higher risk premium for biotech
        
        return min(max(cost_of_equity, 8.0), 20.0)  # Cap between 8% and 20%

def main():
    st.markdown("""
    <div class="dcf-header">
        <h1>📊 DCF Model Builder</h1>
        <p style="font-size: 1.3rem;">Professional Discounted Cash Flow Valuation Model</p>
        <div style="color: #e0e7ff; font-weight: 500;">Build custom DCF models with scenario & sensitivity analysis</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize DCF model
    if 'dcf_model' not in st.session_state:
        st.session_state.dcf_model = DCFModel()
    
    # Sidebar for company selection
    with st.sidebar:
        st.markdown("### 🏢 Company Data Loader")
        
        st.markdown("**📊 Enter a stock ticker to auto-load financial data:**")
        
        # Popular healthcare tickers as examples
        st.markdown("*Popular tickers: MRNA, PFE, JNJ, LLY, ABBV, REGN, VRTX, GILD*")
        
        ticker_input = st.text_input(
            "Stock Ticker:", 
            placeholder="e.g., MRNA",
            help="Enter any publicly traded stock ticker"
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 Load Data", disabled=not ticker_input):
                if ticker_input:
                    load_company_data(ticker_input.upper())
        
        with col2:
            if st.button("🔄 Clear", help="Clear loaded data"):
                # Clear company data from session state
                for key in ['company_data', 'company_ticker', 'dcf_assumptions', 'dcf_projections', 'dcf_valuation']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.rerun()
        
        # Show loaded company status
        if 'company_data' in st.session_state:
            st.success(f"✅ **{st.session_state.company_data['company_info']['name']}** loaded")
            st.caption(f"Ticker: {st.session_state.company_data['company_info']['ticker']}")
        else:
            st.info("💡 Load company data for auto-populated DCF")
        
        st.markdown("---")
        
        st.markdown("### 🎯 How This Works")
        st.markdown("""
        **1. Load Company** 📊
        - Auto-fetches real financial data
        - Calculates historical growth rates
        - Estimates WACC and ratios
        
        **2. Set Assumptions** 🔮
        - Only input forward-looking assumptions
        - Growth rates, margin targets, WACC
        - Historical data auto-populated
        
        **3. Generate DCF** 📈
        - Professional 10-year model
        - Scenario & sensitivity analysis
        - Export to formatted Excel
        """)
        
        st.markdown("---")
        st.markdown("### 💡 DCF Best Practices")
        st.markdown("""
        - **Be Conservative**: Realistic growth assumptions
        - **Sanity Check**: Compare to peer multiples  
        - **Terminal Value**: Should be 60-80% of total
        - **WACC Sensitivity**: Small changes = big impact
        - **Healthcare Focus**: Consider R&D, patents, regulation
        """)
    
    # Show company data preview if just loaded
    show_company_data_preview()
    
    # Main DCF interface
    col1, col2 = st.columns([1, 1])
    
    with col1:
        show_assumptions_input()
    
    with col2:
        show_dcf_results()
    
    # Additional analysis sections
    st.markdown("---")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📈 Scenario Analysis", "🎯 Sensitivity Analysis", "📊 Export Model", "📚 DCF Guide"])
    
    with tab1:
        show_scenario_analysis()
    
    with tab2:
        show_sensitivity_analysis()
    
    with tab3:
        show_export_options()
    
    with tab4:
        show_dcf_guide()

def load_company_data(ticker):
    """Enhanced company data loading with comprehensive financial analysis"""
    
    if not ticker:
        st.error("Please enter a stock ticker")
        return
    
    with st.spinner(f"🔍 Fetching comprehensive data for {ticker}..."):
        scraper = FinancialDataScraper()
        company_data = scraper.get_comprehensive_data(ticker)
    
    if company_data:
        # Store in session state
        st.session_state.company_data = company_data
        st.session_state.company_ticker = ticker
        
        # Simple success message for sidebar
        st.success(f"✅ Loaded {company_data['company_info']['name']}")
        
        # Store flag to show detailed preview in main area
        st.session_state.show_company_preview = True
        
        return True
    else:
        st.error(f"❌ Could not load data for {ticker}")
        return False

def show_company_data_preview():
    """Show detailed company data preview in main area"""
    
    if 'company_data' in st.session_state and st.session_state.get('show_company_preview', False):
        company_data = st.session_state.company_data
        
        # Show key metrics preview
        with st.expander("📊 Key Financial Metrics Preview", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("Revenue (TTM)", f"${company_data['financial_metrics']['revenue_ttm']:.0f}M")
                st.metric("EBITDA Margin", f"{company_data['financial_metrics']['ebitda_margin']:.1f}%")
            
            with col2:
                st.metric("Free Cash Flow", f"${company_data['financial_metrics']['free_cf_ttm']:.0f}M")
                st.metric("Revenue CAGR (3Y)", f"{company_data['growth_metrics']['revenue_cagr_3y']:.1f}%")
            
            with col3:
                st.metric("Market Cap", f"${company_data['market_data']['market_cap']:.0f}M")
                st.metric("Current Price", f"${company_data['market_data']['current_price']:.2f}")
            
            with col4:
                st.metric("Net Debt", f"${company_data['balance_sheet']['net_debt']:.0f}M")
                st.metric("Estimated WACC", f"{company_data['growth_metrics']['estimated_wacc']:.1f}%")
        
        # Clear the preview flag
        st.session_state.show_company_preview = False

def show_assumptions_input():
    """Show simplified DCF assumptions input interface using scraped data"""
    
    st.markdown("### 📝 DCF Assumptions")
    
    # Initialize default values
    revenue_growth_y1 = 10
    revenue_growth_y5 = 5
    revenue_growth_terminal = 2.5
    base_ebitda_margin = 20.0
    target_ebitda_margin = 25.0
    wacc = 10.0
    tax_rate = 21
    capex_revenue = 4.0
    da_revenue = 3.0
    nwc_revenue = 1.0
    
    # Check if we have company data loaded
    if 'company_data' not in st.session_state:
        st.info("👆 Please load company data first to auto-populate financial metrics")
        
        # Manual input fallback
        with st.expander("⚙️ Manual Input (if no company loaded)", expanded=False):
            col1, col2 = st.columns(2)
            
            with col1:
                company_name = st.text_input("Company Name", value="")
                base_revenue = st.number_input("Base Year Revenue ($M)", value=1000.0, step=50.0)
                current_stock_price = st.number_input("Current Stock Price ($)", value=100.0, step=1.0)
            
            with col2:
                shares_outstanding = st.number_input("Shares Outstanding (M)", value=100.0, step=1.0)
                base_ebitda_margin_input = st.number_input("Current EBITDA Margin (%)", value=20.0, step=1.0)
                estimated_wacc_input = st.number_input("Estimated WACC (%)", value=10.0, step=0.5)
                net_debt = st.number_input("Net Debt ($M)", value=0.0, step=50.0)
            
            # Update default values with manual inputs
            base_ebitda_margin = base_ebitda_margin_input
            wacc = estimated_wacc_input
        
        # Store manual assumptions
        if 'dcf_assumptions' not in st.session_state:
            st.session_state.dcf_assumptions = {
                'company_name': company_name if 'company_name' in locals() else '',
                'base_revenue': base_revenue if 'base_revenue' in locals() else 1000.0,
                'current_stock_price': current_stock_price if 'current_stock_price' in locals() else 100.0,
                'shares_outstanding': shares_outstanding if 'shares_outstanding' in locals() else 100.0,
                'base_ebitda_margin': base_ebitda_margin,
                'wacc': wacc,
                'net_debt': net_debt if 'net_debt' in locals() else 0.0
            }
    
    else:
        # Use loaded company data
        company_data = st.session_state.company_data
        
        # Display current company info
        st.markdown(f"""
        <div class="assumption-card">
            <h4>🏢 {company_data['company_info']['name']} ({company_data['company_info']['ticker']})</h4>
            <p><strong>Industry:</strong> {company_data['company_info']['industry']} | <strong>Sector:</strong> {company_data['company_info']['sector']}</p>
        </div>
        """, unsafe_allow_html=True)
        
        # Update defaults with company data
        base_ebitda_margin = company_data['financial_metrics'].get('ebitda_margin', 20.0)
        wacc = company_data['growth_metrics'].get('estimated_wacc', 10.0)
        historical_growth = company_data['growth_metrics'].get('revenue_cagr_3y', 10)
        capex_revenue = company_data['financial_metrics'].get('capex_revenue_ratio', 4.0)
    
    # Forward-looking assumptions (the only inputs user needs to provide)
    st.markdown("#### 🔮 Forward-Looking Assumptions (Your Predictions)")
    
    with st.expander("📈 Revenue Growth Assumptions", expanded=True):
        st.markdown("*Based on your analysis of market opportunities, competition, and company strategy*")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Use historical growth as default suggestion
            historical_growth = st.session_state.get('company_data', {}).get('growth_metrics', {}).get('revenue_cagr_3y', 10)
            suggested_y1 = max(min(historical_growth, 50), -20)  # Cap between -20% and 50%
            
            revenue_growth_y1 = st.slider(
                "Year 1 Revenue Growth (%)", 
                -20, 50, 
                int(suggested_y1), 
                help="Your forecast for next year's revenue growth"
            )
        
        with col2:
            suggested_y5 = max(min(historical_growth * 0.6, 30), -10)  # Typically lower than Y1
            revenue_growth_y5 = st.slider(
                "Year 5 Revenue Growth (%)", 
                -10, 30, 
                int(suggested_y5),
                help="Long-term sustainable growth rate"
            )
        
        with col3:
            revenue_growth_terminal = st.slider(
                "Terminal Growth Rate (%)", 
                0.0, 4.0, 
                2.5, 
                step=0.5,
                help="Long-term GDP growth rate (typically 2-3%)"
            )
    
    with st.expander("💰 Profitability Assumptions", expanded=True):
        st.markdown("*Your view on margin expansion/compression over time*")
        
        col1, col2 = st.columns(2)
        
        with col1:
            base_ebitda_margin = st.slider(
                "Current EBITDA Margin (%)", 
                0, 60, 
                int(base_ebitda_margin), 
                help="Starting point (auto-filled from company data)",
                disabled=True if 'company_data' in st.session_state else False
            )
        
        with col2:
            # Suggest modest improvement
            suggested_target = min(base_ebitda_margin + 3, 50)
            target_ebitda_margin = st.slider(
                "Target EBITDA Margin (%)", 
                0, 60, 
                int(suggested_target),
                help="Your forecast for mature/optimized margin"
            )
    
    with st.expander("🎯 Valuation Parameters", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            wacc = st.slider(
                "WACC - Weighted Average Cost of Capital (%)", 
                6.0, 18.0, 
                wacc, 
                step=0.5,
                help="Discount rate (auto-estimated based on beta and company risk)"
            )
        
        with col2:
            # Tax rate - standard corporate rate
            tax_rate = st.slider("Tax Rate (%)", 15, 35, 21, help="Corporate tax rate")
    
    # Advanced assumptions (optional fine-tuning)
    with st.expander("⚙️ Advanced Assumptions (Optional)", expanded=False):
        st.markdown("*Fine-tune these parameters if needed (defaults are reasonable)*")
        
        col1, col2 = st.columns(2)
        
        with col1:
            capex_revenue = st.slider(
                "Capex as % of Revenue (%)", 
                0.0, 15.0, 
                capex_revenue, 
                step=0.5
            )
            
            da_revenue = st.slider("D&A as % of Revenue (%)", 1.0, 8.0, 3.0, step=0.5)
        
        with col2:
            nwc_revenue = st.slider("Working Capital Impact (% of Revenue)", -3.0, 10.0, 1.0, step=0.5)
    
    # Auto-populate financial data from scraped information
    if 'company_data' in st.session_state:
        company_data = st.session_state.company_data
        
        # Store assumptions combining scraped data with user inputs
        st.session_state.dcf_assumptions = {
            'company_name': company_data['company_info']['name'],
            'ticker': company_data['company_info']['ticker'],
            'base_revenue': company_data['financial_metrics']['revenue_ttm'],
            'current_stock_price': company_data['market_data']['current_price'],
            'shares_outstanding': company_data['market_data']['shares_outstanding'],
            'revenue_growth_y1': revenue_growth_y1,
            'revenue_growth_y5': revenue_growth_y5,
            'revenue_growth_terminal': revenue_growth_terminal,
            'base_ebitda_margin': base_ebitda_margin,
            'target_ebitda_margin': target_ebitda_margin,
            'da_revenue': da_revenue,
            'tax_rate': tax_rate,
            'capex_revenue': capex_revenue,
            'nwc_revenue': nwc_revenue,
            'wacc': wacc,
            'net_debt': company_data['balance_sheet']['net_debt'],
            'historical_data': {
                'revenue_cagr_3y': company_data['growth_metrics']['revenue_cagr_3y'],
                'revenue_cagr_5y': company_data['growth_metrics']['revenue_cagr_5y'],
                'free_cf_ttm': company_data['financial_metrics']['free_cf_ttm']
            }
        }
    else:
        # Manual input mode
        st.session_state.dcf_assumptions = {
            'company_name': company_name if 'company_name' in locals() else 'Manual Input',
            'ticker': 'MANUAL',
            'base_revenue': base_revenue if 'base_revenue' in locals() else 1000.0,
            'current_stock_price': current_stock_price if 'current_stock_price' in locals() else 100.0,
            'shares_outstanding': shares_outstanding if 'shares_outstanding' in locals() else 100.0,
            'revenue_growth_y1': revenue_growth_y1,
            'revenue_growth_y5': revenue_growth_y5,
            'revenue_growth_terminal': revenue_growth_terminal,
            'base_ebitda_margin': base_ebitda_margin,
            'target_ebitda_margin': target_ebitda_margin,
            'da_revenue': da_revenue,
            'tax_rate': tax_rate,
            'capex_revenue': capex_revenue,
            'nwc_revenue': nwc_revenue,
            'wacc': wacc,
            'net_debt': net_debt if 'net_debt' in locals() else 0.0,
        }
    
    # Show assumption summary
    if 'dcf_assumptions' in st.session_state:
        with st.expander("📋 Assumption Summary", expanded=False):
            assumptions = st.session_state.dcf_assumptions
            
            summary_data = {
                'Parameter': [
                    'Company',
                    'Base Revenue ($M)',
                    'Year 1 Growth (%)',
                    'Terminal Growth (%)',
                    'Current EBITDA Margin (%)',
                    'Target EBITDA Margin (%)',
                    'WACC (%)',
                    'Current Stock Price ($)'
                ],
                'Value': [
                    assumptions.get('company_name', 'N/A'),
                    f"{assumptions.get('base_revenue', 0):.0f}",
                    f"{assumptions.get('revenue_growth_y1', 0):.1f}",
                    f"{assumptions.get('revenue_growth_terminal', 0):.1f}",
                    f"{assumptions.get('base_ebitda_margin', 0):.1f}",
                    f"{assumptions.get('target_ebitda_margin', 0):.1f}",
                    f"{assumptions.get('wacc', 0):.1f}",
                    f"{assumptions.get('current_stock_price', 0):.2f}"
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

def show_dcf_results():
    """Show DCF calculation results"""
    
    st.markdown("### 🎯 Valuation Results")
    
    if 'dcf_assumptions' in st.session_state:
        assumptions = st.session_state.dcf_assumptions
        
        # Calculate DCF
        with st.spinner("📊 Running DCF calculations..."):
            projections, valuation = st.session_state.dcf_model.calculate_dcf_valuation(assumptions)
        
        # Display key results
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown(f"""
            <div class="valuation-result">
                <h3>💰 Intrinsic Value</h3>
                <div class="metric-highlight">${valuation['value_per_share']:.2f}</div>
                <p>per share</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            upside_color = "#00C851" if valuation['upside_downside'] > 0 else "#ff4444"
            upside_text = "Upside" if valuation['upside_downside'] > 0 else "Downside"
            
            st.markdown(f"""
            <div class="valuation-result" style="background: linear-gradient(135deg, {upside_color} 0%, {upside_color}CC 100%);">
                <h3>📈 {upside_text}</h3>
                <div class="metric-highlight">{valuation['upside_downside']:+.1f}%</div>
                <p>vs current price ${assumptions['current_stock_price']:.2f}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Valuation breakdown
        st.markdown("#### 📊 Valuation Breakdown")
        
        breakdown_data = {
            'Component': ['PV of Projection Period', 'PV of Terminal Value', 'Enterprise Value', 'Less: Net Debt', 'Equity Value'],
            'Value ($M)': [
                valuation['pv_projection_period'],
                valuation['pv_terminal_value'],
                valuation['enterprise_value'],
                -assumptions['net_debt'],
                valuation['equity_value']
            ],
            '% of EV': [
                valuation['pv_projection_period'] / valuation['enterprise_value'] * 100,
                valuation['pv_terminal_value'] / valuation['enterprise_value'] * 100,
                100,
                -assumptions['net_debt'] / valuation['enterprise_value'] * 100,
                valuation['equity_value'] / valuation['enterprise_value'] * 100
            ]
        }
        
        breakdown_df = pd.DataFrame(breakdown_data)
        st.dataframe(breakdown_df, use_container_width=True)
        
        # Key metrics
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("EV/Revenue (LTM)", f"{valuation['implied_ev_revenue']:.1f}x")
        
        with col2:
            st.metric("Terminal Value %", f"{valuation['pv_terminal_value']/valuation['enterprise_value']*100:.0f}%")
        
        with col3:
            st.metric("Enterprise Value", f"${valuation['enterprise_value']:.0f}M")
        
        # Store results for other tabs
        st.session_state.dcf_projections = projections
        st.session_state.dcf_valuation = valuation
        
        # Financial projections chart
        st.markdown("#### 📈 Financial Projections")
        
        fig = make_subplots(
            rows=2, cols=2,
            subplot_titles=('Revenue & Growth', 'EBITDA & Margin', 'Free Cash Flow', 'Cumulative PV'),
            specs=[[{"secondary_y": True}, {"secondary_y": True}],
                   [{"secondary_y": False}, {"secondary_y": False}]]
        )
        
        # Revenue and growth
        fig.add_trace(go.Bar(x=projections['Year'], y=projections['Revenue'], name='Revenue ($M)', showlegend=False), row=1, col=1)
        fig.add_trace(go.Scatter(x=projections['Year'], y=projections['Revenue_Growth'], name='Growth %', mode='lines+markers', yaxis='y2', showlegend=False), row=1, col=1)
        
        # EBITDA and margin
        fig.add_trace(go.Bar(x=projections['Year'], y=projections['EBITDA'], name='EBITDA ($M)', showlegend=False), row=1, col=2)
        fig.add_trace(go.Scatter(x=projections['Year'], y=projections['EBITDA_Margin'], name='Margin %', mode='lines+markers', yaxis='y4', showlegend=False), row=1, col=2)
        
        # Free cash flow
        fig.add_trace(go.Bar(x=projections['Year'], y=projections['FCFF'], name='FCFF ($M)', showlegend=False), row=2, col=1)
        
        # Cumulative PV
        cumulative_pv = projections['PV_FCFF'].cumsum()
        fig.add_trace(go.Scatter(x=projections['Year'], y=cumulative_pv, name='Cumulative PV', mode='lines+markers', fill='tonexty', showlegend=False), row=2, col=2)
        
        fig.update_layout(height=600, title_text="DCF Model Projections")
        st.plotly_chart(fig, use_container_width=True)
        
    else:
        st.info("👆 Please input your DCF assumptions to see valuation results")

def show_scenario_analysis():
    """Show scenario analysis results"""
    
    st.markdown("### 📈 Scenario Analysis")
    
    if 'dcf_assumptions' in st.session_state:
        assumptions = st.session_state.dcf_assumptions
        
        with st.spinner("🔄 Running scenario analysis..."):
            scenarios = st.session_state.dcf_model.scenario_analysis(assumptions)
        
        # Scenario results table
        scenario_data = {
            'Scenario': [],
            'Value per Share': [],
            'Enterprise Value': [],
            'Upside/Downside': []
        }
        
        base_value = scenarios['Base Case']['valuation']['value_per_share']
        
        for scenario_name, results in scenarios.items():
            scenario_data['Scenario'].append(scenario_name)
            scenario_data['Value per Share'].append(f"${results['valuation']['value_per_share']:.2f}")
            scenario_data['Enterprise Value'].append(f"${results['valuation']['enterprise_value']:.0f}M")
            
            upside = (results['valuation']['value_per_share'] - assumptions['current_stock_price']) / assumptions['current_stock_price'] * 100
            scenario_data['Upside/Downside'].append(f"{upside:+.1f}%")
        
        scenario_df = pd.DataFrame(scenario_data)
        st.dataframe(scenario_df, use_container_width=True)
        
        # Scenario chart
        values = [scenarios[s]['valuation']['value_per_share'] for s in ['Bear Case', 'Base Case', 'Bull Case']]
        colors = ['#ff4444', '#33b5e5', '#00C851']
        
        fig = go.Figure(data=[
            go.Bar(x=['Bear Case', 'Base Case', 'Bull Case'], y=values, marker_color=colors)
        ])
        
        # Add current price line
        fig.add_hline(y=assumptions['current_stock_price'], line_dash="dash", 
                      annotation_text=f"Current Price: ${assumptions['current_stock_price']:.2f}")
        
        fig.update_layout(
            title="Scenario Analysis - Value per Share",
            yaxis_title="Value per Share ($)",
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    else:
        st.info("👆 Please complete DCF assumptions to run scenario analysis")

def show_sensitivity_analysis():
    """Show sensitivity analysis"""
    
    st.markdown("### 🎯 Sensitivity Analysis")
    
    if 'dcf_assumptions' in st.session_state:
        assumptions = st.session_state.dcf_assumptions
        
        # Sensitivity variable selection
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📊 Select Variables to Test")
            
            test_wacc = st.checkbox("WACC", value=True)
            test_terminal_growth = st.checkbox("Terminal Growth Rate", value=True)
            test_revenue_growth = st.checkbox("Year 1 Revenue Growth", value=True)
            test_ebitda_margin = st.checkbox("Target EBITDA Margin", value=False)
        
        with col2:
            st.markdown("#### ⚙️ Sensitivity Ranges")
            
            wacc_range = st.slider("WACC Range (+/-)", 0.5, 3.0, 1.5, step=0.5)
            terminal_range = st.slider("Terminal Growth Range (+/-)", 0.5, 2.0, 1.0, step=0.5)
            revenue_range = st.slider("Revenue Growth Range (+/-)", 2, 10, 5)
            ebitda_range = st.slider("EBITDA Margin Range (+/-)", 1, 5, 2)
        
        if st.button("🚀 Run Sensitivity Analysis"):
            
            # Build sensitivity variables
            sensitivity_vars = {}
            
            if test_wacc:
                base_wacc = assumptions['wacc']
                sensitivity_vars['wacc'] = np.arange(base_wacc - wacc_range, base_wacc + wacc_range + 0.1, 0.5)
            
            if test_terminal_growth:
                base_terminal = 2.5  # Default terminal growth
                sensitivity_vars['revenue_growth_terminal'] = np.arange(base_terminal - terminal_range, base_terminal + terminal_range + 0.1, 0.5)
            
            if test_revenue_growth:
                base_rev_growth = assumptions['revenue_growth_y1']
                sensitivity_vars['revenue_growth_y1'] = np.arange(base_rev_growth - revenue_range, base_rev_growth + revenue_range + 1, 2)
            
            if test_ebitda_margin:
                base_ebitda = assumptions['target_ebitda_margin']
                sensitivity_vars['target_ebitda_margin'] = np.arange(base_ebitda - ebitda_range, base_ebitda + ebitda_range + 1, 1)
            
            with st.spinner("🔄 Running sensitivity analysis..."):
                sensitivity_results = st.session_state.dcf_model.sensitivity_analysis(assumptions, sensitivity_vars)
            
            # Display results
            for var_name, results_df in sensitivity_results.items():
                st.markdown(f"#### {var_name.replace('_', ' ').title()}")
                
                # Chart
                fig = go.Figure()
                
                fig.add_trace(go.Scatter(
                    x=results_df['variable_value'],
                    y=results_df['value_per_share'],
                    mode='lines+markers',
                    name='Value per Share'
                ))
                
                # Add current price line
                fig.add_hline(y=assumptions['current_stock_price'], line_dash="dash",
                             annotation_text=f"Current Price: ${assumptions['current_stock_price']:.2f}")
                
                fig.update_layout(
                    xaxis_title=var_name.replace('_', ' ').title(),
                    yaxis_title="Value per Share ($)",
                    height=300
                )
                
                st.plotly_chart(fig, use_container_width=True)
                
                # Table
                st.dataframe(results_df, use_container_width=True)
    
    else:
        st.info("👆 Please complete DCF assumptions to run sensitivity analysis")

def show_export_options():
    """Show export options for the DCF model"""
    
    st.markdown("### 📊 Export DCF Model")
    
    if 'dcf_projections' in st.session_state and 'dcf_valuation' in st.session_state:
        projections = st.session_state.dcf_projections
        valuation = st.session_state.dcf_valuation
        assumptions = st.session_state.dcf_assumptions
        
        # Excel export option
        st.markdown("#### 📈 Download Excel Model")
        
        if st.button("🔽 Generate Excel DCF Model"):
            excel_buffer = create_excel_dcf_model(projections, valuation, assumptions)
            
            st.download_button(
                label="📊 Download DCF Model.xlsx",
                data=excel_buffer,
                file_name=f"DCF_Model_{assumptions.get('company_name', 'Company').replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # CSV export
        st.markdown("#### 📋 Download CSV Data")
        
        col1, col2 = st.columns(2)
        
        with col1:
            csv_projections = projections.to_csv(index=False)
            st.download_button(
                label="📈 Download Projections CSV",
                data=csv_projections,
                file_name="dcf_projections.csv",
                mime="text/csv"
            )
        
        with col2:
            # Create valuation summary
            valuation_summary = pd.DataFrame([
                ['Enterprise Value', f"${valuation['enterprise_value']:.0f}M"],
                ['Equity Value', f"${valuation['equity_value']:.0f}M"],
                ['Value per Share', f"${valuation['value_per_share']:.2f}"],
                ['Current Price', f"${assumptions['current_stock_price']:.2f}"],
                ['Upside/Downside', f"{valuation['upside_downside']:+.1f}%"]
            ], columns=['Metric', 'Value'])
            
            csv_valuation = valuation_summary.to_csv(index=False)
            st.download_button(
                label="💰 Download Valuation CSV",
                data=csv_valuation,
                file_name="dcf_valuation.csv",
                mime="text/csv"
            )
        
        # Model summary
        st.markdown("#### 📋 Model Summary")
        
        summary_data = {
            'Parameter': ['Company', 'Base Revenue', 'WACC', 'Terminal Growth', 'Value per Share', 'Upside/Downside'],
            'Value': [
                assumptions.get('company_name', 'N/A'),
                f"${assumptions['base_revenue']:.0f}M",
                f"{assumptions['wacc']:.1f}%",
                f"{assumptions['revenue_growth_terminal']:.1f}%",
                f"${valuation['value_per_share']:.2f}",
                f"{valuation['upside_downside']:+.1f}%"
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        st.dataframe(summary_df, use_container_width=True)
    
    else:
        st.info("👆 Please complete the DCF model to access export options")

def create_excel_dcf_model(projections, valuation, assumptions):
    """Create professionally formatted Excel DCF model following investment banking standards"""
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create workbook
    wb = Workbook()
    
    # Define professional styling
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    subheader_font = Font(name='Arial', size=11, bold=True)
    subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    number_font = Font(name='Arial', size=10)
    currency_font = Font(name='Arial', size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal='center', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    
    # 1. Summary Sheet (Main DCF Model)
    ws_summary = wb.active
    ws_summary.title = "DCF Summary"
    
    # Company header
    ws_summary['A1'] = f"DCF Valuation Model - {assumptions.get('company_name', 'Company')}"
    ws_summary['A1'].font = Font(name='Arial', size=16, bold=True)
    ws_summary.merge_cells('A1:K1')
    
    ws_summary['A2'] = f"Ticker: {assumptions.get('ticker', 'N/A')}"
    ws_summary['A2'].font = Font(name='Arial', size=12)
    
    ws_summary['A3'] = f"Valuation Date: {datetime.now().strftime('%B %d, %Y')}"
    ws_summary['A3'].font = Font(name='Arial', size=11)
    
    # Key metrics summary
    row = 5
    ws_summary[f'A{row}'] = "KEY VALUATION METRICS"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    metrics_data = [
        ['Enterprise Value ($M)', f"{valuation['enterprise_value']:.0f}"],
        ['Equity Value ($M)', f"{valuation['equity_value']:.0f}"],
        ['Value per Share ($)', f"{valuation['value_per_share']:.2f}"],
        ['Current Stock Price ($)', f"{assumptions['current_stock_price']:.2f}"],
        ['Upside/(Downside)', f"{valuation['upside_downside']:+.1f}%"],
        ['Implied EV/Revenue (LTM)', f"{valuation.get('implied_ev_revenue', 0):.1f}x"]
    ]
    
    for i, (metric, value) in enumerate(metrics_data):
        row_idx = row + 1 + i
        ws_summary[f'A{row_idx}'] = metric
        ws_summary[f'B{row_idx}'] = value
        ws_summary[f'A{row_idx}'].font = number_font
        ws_summary[f'B{row_idx}'].font = currency_font
        ws_summary[f'B{row_idx}'].alignment = right_alignment
    
    # Financial projections table
    row = 14
    ws_summary[f'A{row}'] = "FINANCIAL PROJECTIONS ($M)"
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary.merge_cells(f'A{row}:K{row}')
    
    # Headers
    headers = ['', 'Base Year'] + [f'Year {i}' for i in range(1, 11)] + ['Terminal']
    for col_idx, header in enumerate(headers):
        cell = ws_summary.cell(row=row+1, column=col_idx+1, value=header)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Revenue projections
    revenue_row = row + 2
    ws_summary[f'A{revenue_row}'] = "Revenue"
    ws_summary[f'A{revenue_row}'].font = subheader_font
    
    # Base year revenue
    ws_summary[f'B{revenue_row}'] = assumptions['base_revenue']
    ws_summary[f'B{revenue_row}'].number_format = '#,##0'
    
    # Projected revenues
    for i, revenue in enumerate(projections['Revenue']):
        col_letter = chr(67 + i)  # Starting from column C
        ws_summary[f'{col_letter}{revenue_row}'] = revenue
        ws_summary[f'{col_letter}{revenue_row}'].number_format = '#,##0'
    
    # Revenue growth rates
    growth_row = revenue_row + 1
    ws_summary[f'A{growth_row}'] = "Revenue Growth %"
    ws_summary[f'A{growth_row}'].font = number_font
    
    for i, growth in enumerate(projections['Revenue_Growth']):
        col_letter = chr(67 + i)  # Starting from column C
        ws_summary[f'{col_letter}{growth_row}'] = growth / 100
        ws_summary[f'{col_letter}{growth_row}'].number_format = '0.0%'
    
    # Terminal growth
    terminal_col = chr(67 + len(projections))
    ws_summary[f'{terminal_col}{growth_row}'] = assumptions['revenue_growth_terminal'] / 100
    ws_summary[f'{terminal_col}{growth_row}'].number_format = '0.0%'
    
    # EBITDA projections
    ebitda_row = growth_row + 2
    ws_summary[f'A{ebitda_row}'] = "EBITDA"
    ws_summary[f'A{ebitda_row}'].font = subheader_font
    
    for i, ebitda in enumerate(projections['EBITDA']):
        col_letter = chr(67 + i)
        ws_summary[f'{col_letter}{ebitda_row}'] = ebitda
        ws_summary[f'{col_letter}{ebitda_row}'].number_format = '#,##0'
    
    # EBITDA margins
    margin_row = ebitda_row + 1
    ws_summary[f'A{margin_row}'] = "EBITDA Margin %"
    ws_summary[f'A{margin_row}'].font = number_font
    
    for i, margin in enumerate(projections['EBITDA_Margin']):
        col_letter = chr(67 + i)
        ws_summary[f'{col_letter}{margin_row}'] = margin / 100
        ws_summary[f'{col_letter}{margin_row}'].number_format = '0.0%'
    
    # Free Cash Flow
    fcf_row = margin_row + 2
    ws_summary[f'A{fcf_row}'] = "Free Cash Flow"
    ws_summary[f'A{fcf_row}'].font = subheader_font
    
    for i, fcf in enumerate(projections['FCFF']):
        col_letter = chr(67 + i)
        ws_summary[f'{col_letter}{fcf_row}'] = fcf
        ws_summary[f'{col_letter}{fcf_row}'].number_format = '#,##0'
    
    # Present Value of FCF
    pv_row = fcf_row + 1
    ws_summary[f'A{pv_row}'] = "Present Value of FCF"
    ws_summary[f'A{pv_row}'].font = number_font
    
    for i, pv_fcf in enumerate(projections['PV_FCFF']):
        col_letter = chr(67 + i)
        ws_summary[f'{col_letter}{pv_row}'] = pv_fcf
        ws_summary[f'{col_letter}{pv_row}'].number_format = '#,##0'
    
    # Valuation summary
    val_row = pv_row + 3
    ws_summary[f'A{val_row}'] = "VALUATION SUMMARY"
    ws_summary[f'A{val_row}'].font = header_font
    ws_summary[f'A{val_row}'].fill = header_fill
    ws_summary.merge_cells(f'A{val_row}:C{val_row}')
    
    val_data = [
        ['PV of Projection Period', valuation['pv_projection_period']],
        ['PV of Terminal Value', valuation['pv_terminal_value']],
        ['Enterprise Value', valuation['enterprise_value']],
        ['Less: Net Debt', -assumptions['net_debt']],
        ['Equity Value', valuation['equity_value']],
        ['Shares Outstanding (M)', assumptions['shares_outstanding']],
        ['Value per Share', valuation['value_per_share']]
    ]
    
    for i, (item, value) in enumerate(val_data):
        row_idx = val_row + 1 + i
        ws_summary[f'A{row_idx}'] = item
        ws_summary[f'B{row_idx}'] = value
        ws_summary[f'A{row_idx}'].font = number_font
        ws_summary[f'B{row_idx}'].font = currency_font
        ws_summary[f'B{row_idx}'].alignment = right_alignment
        
        if item == 'Value per Share':
            ws_summary[f'B{row_idx}'].number_format = '$#,##0.00'
        else:
            ws_summary[f'B{row_idx}'].number_format = '#,##0'
    
    # 2. Assumptions Sheet
    ws_assumptions = wb.create_sheet("Assumptions")
    
    ws_assumptions['A1'] = "DCF MODEL ASSUMPTIONS"
    ws_assumptions['A1'].font = header_font
    ws_assumptions['A1'].fill = header_fill
    ws_assumptions.merge_cells('A1:C1')
    
    # Growth assumptions
    assumptions_data = [
        ['GROWTH ASSUMPTIONS', '', ''],
        ['Year 1 Revenue Growth (%)', assumptions['revenue_growth_y1'], '%'],
        ['Year 5 Revenue Growth (%)', assumptions['revenue_growth_y5'], '%'],
        ['Terminal Growth Rate (%)', assumptions['revenue_growth_terminal'], '%'],
        ['', '', ''],
        ['PROFITABILITY ASSUMPTIONS', '', ''],
        ['Base EBITDA Margin (%)', assumptions['base_ebitda_margin'], '%'],
        ['Target EBITDA Margin (%)', assumptions['target_ebitda_margin'], '%'],
        ['Tax Rate (%)', assumptions['tax_rate'], '%'],
        ['', '', ''],
        ['INVESTMENT ASSUMPTIONS', '', ''],
        ['Capex as % of Revenue (%)', assumptions['capex_revenue'], '%'],
        ['D&A as % of Revenue (%)', assumptions['da_revenue'], '%'],
        ['NWC as % of Revenue (%)', assumptions['nwc_revenue'], '%'],
        ['', '', ''],
        ['VALUATION ASSUMPTIONS', '', ''],
        ['WACC (%)', assumptions['wacc'], '%'],
        ['Net Debt ($M)', assumptions['net_debt'], '$M'],
        ['Shares Outstanding (M)', assumptions['shares_outstanding'], 'M']
    ]
    
    for row_idx, (param, value, unit) in enumerate(assumptions_data, start=3):
        ws_assumptions[f'A{row_idx}'] = param
        ws_assumptions[f'B{row_idx}'] = value
        ws_assumptions[f'C{row_idx}'] = unit
        
        if param in ['GROWTH ASSUMPTIONS', 'PROFITABILITY ASSUMPTIONS', 'INVESTMENT ASSUMPTIONS', 'VALUATION ASSUMPTIONS']:
            ws_assumptions[f'A{row_idx}'].font = subheader_font
            ws_assumptions[f'A{row_idx}'].fill = subheader_fill
        else:
            ws_assumptions[f'A{row_idx}'].font = number_font
    
    # 3. Detailed Projections Sheet
    ws_projections = wb.create_sheet("Detailed Projections")
    
    # Add projections dataframe
    for r in dataframe_to_rows(projections, index=False, header=True):
        ws_projections.append(r)
    
    # Format headers
    for cell in ws_projections[1]:
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_assumptions, ws_projections]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def show_dcf_guide():
    """Show DCF methodology guide"""
    
    st.markdown("### 📚 DCF Methodology Guide")
    
    with st.expander("🎯 What is DCF?", expanded=True):
        st.markdown("""
        **Discounted Cash Flow (DCF)** is a valuation method that estimates the value of a company based on its projected future cash flows, discounted back to present value.
        
        **Key Formula**: `Enterprise Value = Σ(FCF / (1 + WACC)^t) + Terminal Value / (1 + WACC)^n`
        
        **Why DCF?**
        - **Intrinsic Value**: Based on fundamental business performance
        - **Forward-Looking**: Uses future projections, not just historical data
        - **Comprehensive**: Considers all aspects of business performance
        - **Flexible**: Can model different scenarios and assumptions
        """)
    
    with st.expander("📊 Key Components"):
        st.markdown("""
        **1. Free Cash Flow to the Firm (FCFF)**
        ```
        FCFF = EBIT × (1 - Tax Rate) + Depreciation - CapEx - Change in NWC
        ```
        
        **2. Weighted Average Cost of Capital (WACC)**
        ```
        WACC = (E/V × Re) + (D/V × Rd × (1 - Tax Rate))
        ```
        
        **3. Terminal Value**
        ```
        Terminal Value = FCF(final year) × (1 + g) / (WACC - g)
        ```
        
        **4. Enterprise to Equity Value**
        ```
        Equity Value = Enterprise Value - Net Debt
        Value per Share = Equity Value / Shares Outstanding
        ```
        """)
    
    with st.expander("⚖️ Key Assumptions"):
        st.markdown("""
        **Revenue Growth**: 
        - Year 1-5: Based on market size, competition, product lifecycle
        - Terminal: Usually 2-4%, not exceeding long-term GDP growth
        
        **Margins**:
        - Consider operating leverage, competition, industry dynamics
        - Healthcare companies often have high R&D costs but strong margins
        
        **WACC**:
        - Risk-free rate + equity risk premium × beta + debt costs
        - Healthcare: typically 8-12% depending on company stage/risk
        
        **Terminal Growth**:
        - Conservative: 2-3% (long-term GDP/inflation)
        - Should not exceed long-term economic growth
        """)
    
    with st.expander("🎯 Healthcare-Specific Considerations"):
        st.markdown("""
        **Biotech & Pharma Valuation**:
        - **Pipeline Risk**: Weight by probability of success
        - **Patent Cliffs**: Model revenue decline when patents expire
        - **R&D Intensity**: High upfront costs, long development cycles
        - **Regulatory Risk**: FDA approval uncertainty
        
        **Key Metrics**:
        - **Risk-Adjusted NPV**: Probability-weight pipeline assets
        - **Peak Sales**: Estimate maximum drug revenue potential
        - **Time to Market**: Consider development timelines
        - **Competition**: Model market share erosion
        
        **Typical Ranges**:
        - **WACC**: 9-15% (higher for early-stage biotech)
        - **Terminal Growth**: 2-4%
        - **EBITDA Margins**: 20-40% for mature pharma, negative for early biotech
        """)
    
    with st.expander("⚠️ Common Pitfalls"):
        st.markdown("""
        **1. Over-Optimistic Growth**
        - Reality check against industry/market size
        - Consider competitive dynamics
        
        **2. Terminal Value Dominance**
        - Should be 60-80% of total value
        - If >90%, projections may be too conservative
        
        **3. Ignoring Cyclicality**
        - Normalize for economic cycles
        - Use multiple scenarios
        
        **4. WACC Assumptions**
        - Small changes have huge impact
        - Regularly update based on market conditions
        
        **5. Working Capital**
        - Don't forget seasonal effects
        - Model realistic collection/payment cycles
        """)

if __name__ == "__main__":
    main() 